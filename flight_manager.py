#!/usr/bin/env python3
"""
نظام إدارة رحلات مطار مسقط الدولي - MCT
=====================================

يجمع ويدير الرحلات المغادرة من مسقط بنظام ذكي:
- ملف مباشر: الرحلات خلال 24 ساعة القادمة
- أرشيف دائم: جميع الرحلات التاريخية بدون تكرار
- تحديث ذكي: يحدث البيانات المتغيرة فقط

المتطلبات:
    pip install requests beautifulsoup4 cloudscraper openpyxl pandas
    pip install playwright && playwright install chromium  # اختياري
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
import time
import random
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any, Iterable

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("⚠️  openpyxl غير مثبت - تصدير Excel معطل. قم بتثبيته: pip install openpyxl")

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)

# ══════════════════════════════════════════════════════════════════════════════
# الإعدادات
# ══════════════════════════════════════════════════════════════════════════════

BASE_URL = "https://www.muscatairport.co.om"
SOURCE_URL = "https://www.muscatairport.co.om/flightstatusframe?type=2"

# مسارات الملفات
OUTPUT_DIR = Path("flight_data")
LIVE_JSON = OUTPUT_DIR / "live_flights.json"        # الرحلات الحالية (24 ساعة)
ARCHIVE_JSON = OUTPUT_DIR / "archive_flights.json"  # الأرشيف الكامل
LIVE_EXCEL = OUTPUT_DIR / "live_flights.xlsx"       # Excel للرحلات الحالية
ARCHIVE_EXCEL = OUTPUT_DIR / "archive_flights.xlsx" # Excel للأرشيف
META_JSON = OUTPUT_DIR / "metadata.json"            # معلومات التحديث

LOCAL_TZ = ZoneInfo("Asia/Muscat")
RETRY_DELAY = 3

MONTHS = {m: m for m in ("JAN","FEB","MAR","APR","MAY","JUN",
                          "JUL","AUG","SEP","OCT","NOV","DEC")}

DESTINATION_ALIASES = {
    "LON": "LHR", "LONDON": "LHR", "LONDON HEATHROW": "LHR", "HEATHROW": "LHR",
    "DXB": "DXB", "DUBAI": "DXB",
    "DOH": "DOH", "DOHA": "DOH",
    "JED": "JED", "JEDDAH": "JED",
}

# ══════════════════════════════════════════════════════════════════════════════
# نموذج البيانات
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Flight:
    """نموذج بيانات الرحلة"""
    code: str                    # رقم الرحلة (مثل: WY101)
    date: str                    # التاريخ (مثل: 9MAY)
    destination: str             # رمز الوجهة (مثل: LHR)
    stdEtd: str                  # الوقت المجدول/المتوقع (مثل: 1415/1430)
    status: str = ""             # الحالة (Scheduled, Boarding, Departed...)
    airline: str = ""            # شركة الطيران
    sourceDestination: str = ""  # الوجهة الكاملة (مثل: London Heathrow)
    scheduledAt: str = ""        # الوقت المجدول ISO format
    estimatedAt: str = ""        # الوقت المتوقع ISO format
    gate: str = ""               # البوابة
    terminal: str = ""           # المبنى
    remarks: str = ""            # ملاحظات
    lastUpdated: str = ""        # آخر تحديث
    firstSeen: str = ""          # أول ظهور

# ══════════════════════════════════════════════════════════════════════════════
# دوال مساعدة
# ══════════════════════════════════════════════════════════════════════════════

def now_iso() -> str:
    """الوقت الحالي بصيغة ISO"""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def local_now() -> datetime:
    """الوقت المحلي (مسقط)"""
    return datetime.now(LOCAL_TZ).replace(microsecond=0)

def normalize_space(value: Any) -> str:
    """إزالة المسافات الزائدة"""
    return re.sub(r"\s+", " ", str(value or "")).strip()

def normalize_flight_code(value: str) -> str:
    """توحيد رقم الرحلة"""
    return re.sub(r"\s+", "", normalize_space(value).upper())

def normalize_time(value: str) -> str:
    """تحويل الوقت إلى صيغة موحدة HHMM"""
    s = normalize_space(value)
    if not s:
        return ""
    m = re.search(r"\b(\d{1,2})[:.](\d{2})\b", s)
    if m:
        hh, mm = int(m.group(1)), int(m.group(2))
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}{mm:02d}"
    digits = re.sub(r"\D", "", s)
    if 3 <= len(digits) <= 4:
        digits = digits.zfill(4)
        hh, mm = int(digits[:2]), int(digits[2:4])
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}{mm:02d}"
    return ""

def date_key_from_datetime(dt: datetime | None = None) -> str:
    """تحويل datetime إلى صيغة التاريخ (مثل: 9MAY)"""
    dt = dt or datetime.now()
    return f"{dt.day}{dt.strftime('%b').upper()}"

def normalize_date_key(value: str) -> str:
    """توحيد صيغة التاريخ"""
    s = normalize_space(value).upper()
    if not s:
        return date_key_from_datetime()
    m = re.search(r"\b(\d{1,2})\s*[-/]?\s*([A-Z]{3})\s*(?:[-/]?\s*\d{2,4})?\b", s)
    if m and m.group(2) in MONTHS:
        return f"{int(m.group(1))}{m.group(2)}"
    m = re.search(r"\b(\d{4})-(\d{2})-(\d{2})\b", s)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return f"{dt.day}{dt.strftime('%b').upper()}"
        except ValueError:
            pass
    return date_key_from_datetime()

def normalize_destination(value: str) -> tuple[str, str]:
    """توحيد رمز الوجهة واستخراج الاسم الكامل"""
    raw = normalize_space(value)
    upper = raw.upper()
    if upper in DESTINATION_ALIASES:
        return DESTINATION_ALIASES[upper], raw
    m = re.search(r"\(([A-Z]{3})\)", upper)
    if m:
        code = DESTINATION_ALIASES.get(m.group(1), m.group(1))
        return code, raw
    if re.fullmatch(r"[A-Z]{3}", upper):
        return DESTINATION_ALIASES.get(upper, upper), raw
    return DESTINATION_ALIASES.get(upper, raw), raw

def build_std_etd(scheduled: str, estimated: str) -> str:
    """بناء حقل الوقت المجدول/المتوقع"""
    std = normalize_time(scheduled)
    etd = normalize_time(estimated)
    if std and etd:
        return f"{std}/{etd}"
    return std or etd or ""

def parse_airport_datetime(value: str) -> datetime | None:
    """تحليل التاريخ والوقت من نص المطار"""
    s = normalize_space(value)
    if not s:
        return None
    
    patterns = [
        r"\b(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})\b",
        r"\b(\d{1,2})-(\d{1,2})-(\d{4})\s+(\d{1,2}):(\d{2})\b",
        r"\b(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{2})\b",
    ]
    
    for pattern in patterns:
        m = re.search(pattern, s)
        if not m:
            continue
        try:
            if pattern.startswith(r"\b(\d{4})"):
                year, month, day, hour, minute = map(int, m.groups())
            else:
                day, month, year, hour, minute = map(int, m.groups())
            return datetime(year, month, day, hour, minute, tzinfo=LOCAL_TZ)
        except ValueError:
            return None
    
    return None

def iso_or_empty(dt: datetime | None) -> str:
    """تحويل datetime إلى ISO أو نص فارغ"""
    return dt.isoformat() if dt else ""

def flight_identity(flight: dict[str, Any]) -> str:
    """
    مفتاح فريد للرحلة يستخدم للمقارنة والدمج
    المفتاح = رقم الرحلة + التاريخ + الوجهة + الوقت المجدول
    """
    scheduled = normalize_space(flight.get("scheduledAt", ""))
    return "|".join([
        normalize_flight_code(str(flight.get("code", ""))),
        normalize_date_key(str(flight.get("date", ""))),
        normalize_space(flight.get("destination", "")).upper(),
        scheduled,
    ])

# ══════════════════════════════════════════════════════════════════════════════
# جلب البيانات من الموقع
# ══════════════════════════════════════════════════════════════════════════════

def request_with_cloudscraper(url: str) -> str:
    """محاولة الجلب باستخدام cloudscraper"""
    try:
        import cloudscraper
        scraper = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        response = scraper.get(url, timeout=30)
        response.raise_for_status()
        return response.text
    except Exception as e:
        logger.debug(f"cloudscraper فشل: {e}")
        raise

def request_with_standard(url: str) -> str:
    """محاولة الجلب باستخدام requests عادي"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }
    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.text

def request_with_playwright(url: str) -> str:
    """محاولة الجلب باستخدام playwright (متصفح حقيقي)"""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, wait_until="networkidle", timeout=30000)
            html = page.content()
            browser.close()
            return html
    except Exception as e:
        logger.debug(f"playwright فشل: {e}")
        raise

def request_live_html(url: str) -> str:
    """جلب صفحة الموقع بعدة استراتيجيات"""
    strategies = [
        ("cloudscraper", request_with_cloudscraper),
        ("requests", request_with_standard),
        ("playwright", request_with_playwright),
    ]
    
    for name, func in strategies:
        try:
            logger.info(f"🔄 جاري الجلب باستخدام {name}...")
            html = func(url)
            if html and len(html) > 500:
                logger.info(f"✅ نجح الجلب باستخدام {name}")
                return html
        except Exception as e:
            logger.warning(f"❌ {name} فشل: {e}")
            time.sleep(RETRY_DELAY)
    
    raise RuntimeError("فشلت جميع محاولات جلب البيانات")

# ══════════════════════════════════════════════════════════════════════════════
# تحليل البيانات
# ══════════════════════════════════════════════════════════════════════════════

def parse_table_flights(html: str) -> list[Flight]:
    """
    تحليل جدول الرحلات من HTML
    الأعمدة: airline, destination, flight code, scheduled, estimated, social, status
    """
    soup = BeautifulSoup(html, "html.parser")
    flights: list[Flight] = []
    
    table = soup.find("table")
    if not table:
        return []
    
    for row in table.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 7:
            continue
        
        airline = normalize_space(cells[0].get_text(" "))
        dest_raw = normalize_space(cells[1].get_text(" "))
        flight_code = normalize_flight_code(cells[2].get_text(" "))
        sched_raw = normalize_space(cells[3].get_text(" "))
        est_raw = normalize_space(cells[4].get_text(" "))
        status = normalize_space(cells[6].get_text(" "))
        
        # التحقق من صحة رقم الرحلة
        if not re.match(r"^[A-Z0-9]{2,3}\d{1,5}[A-Z]?$", flight_code):
            continue
        
        # استخراج التاريخ
        date_key = date_key_from_datetime()
        dm = re.search(r"(\d{1,2})/(\d{2})/(\d{4})", sched_raw)
        if dm:
            try:
                dt = datetime(int(dm.group(3)), int(dm.group(2)), int(dm.group(1)))
                date_key = f"{dt.day}{dt.strftime('%b').upper()}"
            except ValueError:
                pass
        
        def extract_time(s: str) -> str:
            t = re.search(r"\b(\d{1,2}):(\d{2})\b", s)
            if t:
                return f"{int(t.group(1)):02d}{int(t.group(2)):02d}"
            return ""
        
        dest, source_dest = normalize_destination(dest_raw)
        scheduled_dt = parse_airport_datetime(sched_raw)
        estimated_dt = parse_airport_datetime(est_raw)
        
        flights.append(Flight(
            code=flight_code,
            date=date_key,
            destination=dest,
            stdEtd=build_std_etd(extract_time(sched_raw), extract_time(est_raw)),
            status=status,
            airline=airline,
            sourceDestination=source_dest,
            scheduledAt=iso_or_empty(scheduled_dt),
            estimatedAt=iso_or_empty(estimated_dt),
            lastUpdated=now_iso(),
            firstSeen=now_iso(),
        ))
    
    return dedupe_flights(flights)

def dedupe_flights(flights: list[Flight]) -> list[Flight]:
    """إزالة التكرارات"""
    seen: set[tuple] = set()
    out: list[Flight] = []
    for f in flights:
        key = (f.code, f.date, f.destination, f.stdEtd)
        if key not in seen:
            seen.add(key)
            out.append(f)
    return out

def parse_flights(html: str) -> list[Flight]:
    """تحليل الرحلات من HTML"""
    return parse_table_flights(html)

# ══════════════════════════════════════════════════════════════════════════════
# فلترة وتصنيف الرحلات
# ══════════════════════════════════════════════════════════════════════════════

def filter_next_24_hours(flights: list[Flight], now: datetime | None = None) -> list[Flight]:
    """الاحتفاظ فقط بالرحلات خلال 24 ساعة القادمة"""
    now = now or local_now()
    end = now + timedelta(hours=24)
    out: list[Flight] = []
    
    for flight in flights:
        dt = parse_airport_datetime(flight.scheduledAt) or parse_airport_datetime(flight.estimatedAt)
        if dt and now <= dt <= end:
            out.append(flight)
    
    return out

# ══════════════════════════════════════════════════════════════════════════════
# دمج وتحديث الأرشيف
# ══════════════════════════════════════════════════════════════════════════════

def merge_archive(existing: list[dict[str, Any]], new_flights: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    دمج الرحلات الجديدة مع الأرشيف الحالي
    - إذا وجدت رحلة بنفس المفتاح: تحديث البيانات المتغيرة فقط
    - إذا لم توجد: إضافة الرحلة الجديدة
    - الاحتفاظ بتاريخ أول ظهور (firstSeen)
    """
    archive_map: dict[str, dict[str, Any]] = {}
    
    # بناء خريطة من الأرشيف الحالي
    for flight in existing:
        key = flight_identity(flight)
        archive_map[key] = flight
    
    # معالجة الرحلات الجديدة
    for new_flight in new_flights:
        key = flight_identity(new_flight)
        
        if key in archive_map:
            # تحديث الرحلة الموجودة
            existing_flight = archive_map[key]
            
            # الاحتفاظ بتاريخ أول ظهور
            new_flight["firstSeen"] = existing_flight.get("firstSeen", new_flight.get("firstSeen", now_iso()))
            
            # تحديث الحقول المتغيرة
            new_flight["lastUpdated"] = now_iso()
            
            # استبدال الرحلة القديمة بالجديدة
            archive_map[key] = new_flight
        else:
            # رحلة جديدة تماماً
            new_flight["firstSeen"] = new_flight.get("firstSeen", now_iso())
            new_flight["lastUpdated"] = now_iso()
            archive_map[key] = new_flight
    
    # تحويل الخريطة إلى قائمة مرتبة
    result = list(archive_map.values())
    
    # ترتيب حسب التاريخ المجدول (الأحدث أولاً)
    result.sort(key=lambda x: x.get("scheduledAt", ""), reverse=True)
    
    return result

# ══════════════════════════════════════════════════════════════════════════════
# حفظ وقراءة البيانات
# ══════════════════════════════════════════════════════════════════════════════

def write_json(path: Path, payload: Any) -> None:
    """حفظ البيانات بصيغة JSON"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

def read_json_list(path: Path) -> list[dict[str, Any]]:
    """قراءة البيانات من JSON"""
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []

def write_meta(status: str, message: str, live_count: int, archive_count: int) -> None:
    """حفظ معلومات التحديث"""
    write_json(META_JSON, {
        "status": status,
        "message": message,
        "liveFlightsCount": live_count,
        "archiveFlightsCount": archive_count,
        "sourceUrl": SOURCE_URL,
        "updatedAt": now_iso(),
        "timezone": str(LOCAL_TZ),
    })

# ══════════════════════════════════════════════════════════════════════════════
# تصدير إلى Excel
# ══════════════════════════════════════════════════════════════════════════════

def create_excel(flights: list[dict[str, Any]], output_path: Path, title: str) -> None:
    """إنشاء ملف Excel منسق"""
    if not HAS_EXCEL:
        logger.warning("⚠️  openpyxl غير متوفر - تخطي تصدير Excel")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flights"
    
    # العناوين
    headers = [
        "رقم الرحلة", "التاريخ", "شركة الطيران", "الوجهة", "الوجهة الكاملة",
        "الوقت المجدول", "الوقت المتوقع", "الحالة", "البوابة", "المبنى",
        "ملاحظات", "أول ظهور", "آخر تحديث"
    ]
    
    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # كتابة البيانات
    for row, flight in enumerate(flights, 2):
        ws.cell(row, 1, flight.get("code", ""))
        ws.cell(row, 2, flight.get("date", ""))
        ws.cell(row, 3, flight.get("airline", ""))
        ws.cell(row, 4, flight.get("destination", ""))
        ws.cell(row, 5, flight.get("sourceDestination", ""))
        ws.cell(row, 6, flight.get("stdEtd", "").split("/")[0] if "/" in flight.get("stdEtd", "") else flight.get("stdEtd", ""))
        ws.cell(row, 7, flight.get("stdEtd", "").split("/")[1] if "/" in flight.get("stdEtd", "") else "")
        ws.cell(row, 8, flight.get("status", ""))
        ws.cell(row, 9, flight.get("gate", ""))
        ws.cell(row, 10, flight.get("terminal", ""))
        ws.cell(row, 11, flight.get("remarks", ""))
        
        # تنسيق التواريخ
        first_seen = flight.get("firstSeen", "")
        last_updated = flight.get("lastUpdated", "")
        
        if first_seen:
            try:
                dt = datetime.fromisoformat(first_seen.replace("Z", "+00:00"))
                ws.cell(row, 12, dt.strftime("%Y-%m-%d %H:%M"))
            except:
                ws.cell(row, 12, first_seen)
        
        if last_updated:
            try:
                dt = datetime.fromisoformat(last_updated.replace("Z", "+00:00"))
                ws.cell(row, 13, dt.strftime("%Y-%m-%d %H:%M"))
            except:
                ws.cell(row, 13, last_updated)
    
    # تنسيق الأعمدة
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # حفظ الملف
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"✅ تم حفظ Excel: {output_path}")

# ══════════════════════════════════════════════════════════════════════════════
# الوظيفة الرئيسية
# ══════════════════════════════════════════════════════════════════════════════

def main() -> int:
    """تشغيل النظام الكامل"""
    try:
        logger.info("🚀 بدء نظام إدارة رحلات مطار مسقط")
        logger.info(f"📍 المصدر: {SOURCE_URL}")
        
        # جلب البيانات
        html = request_live_html(SOURCE_URL)
        flights = parse_flights(html)
        
        if not flights:
            logger.warning("⚠️  لم يتم العثور على رحلات - الاحتفاظ بالبيانات الحالية")
            existing_live = read_json_list(LIVE_JSON)
            existing_archive = read_json_list(ARCHIVE_JSON)
            write_meta(
                status="no_flights_parsed",
                message="لم يتم تحليل أي رحلات - الاحتفاظ بالبيانات الحالية",
                live_count=len(existing_live),
                archive_count=len(existing_archive),
            )
            return 0
        
        logger.info(f"✅ تم جلب {len(flights)} رحلة")
        
        # فلترة الرحلات (24 ساعة القادمة)
        live_flights = filter_next_24_hours(flights)
        live_payload = [asdict(f) for f in live_flights]
        
        # جميع الرحلات للأرشيف
        all_flights_payload = [asdict(f) for f in flights]
        
        # دمج مع الأرشيف
        existing_archive = read_json_list(ARCHIVE_JSON)
        archive_payload = merge_archive(existing_archive, all_flights_payload)
        
        # حفظ JSON
        write_json(LIVE_JSON, live_payload)
        write_json(ARCHIVE_JSON, archive_payload)
        
        logger.info(f"💾 الرحلات المباشرة: {len(live_payload)}")
        logger.info(f"📚 الأرشيف الكامل: {len(archive_payload)}")
        
        # تصدير إلى Excel
        if HAS_EXCEL:
            create_excel(live_payload, LIVE_EXCEL, "الرحلات المباشرة")
            create_excel(archive_payload, ARCHIVE_EXCEL, "أرشيف الرحلات")
        
        # حفظ معلومات التحديث
        write_meta(
            status="success",
            message="تم التحديث بنجاح",
            live_count=len(live_payload),
            archive_count=len(archive_payload),
        )
        
        logger.info("✨ اكتمل التحديث بنجاح!")
        logger.info(f"📂 الملفات المحفوظة في: {OUTPUT_DIR.absolute()}")
        
        return 0
        
    except Exception as exc:
        logger.error(f"❌ خطأ: {exc}", exc_info=True)
        
        # الاحتفاظ بالبيانات الحالية في حالة الخطأ
        existing_live = read_json_list(LIVE_JSON)
        existing_archive = read_json_list(ARCHIVE_JSON)
        
        write_meta(
            status="error",
            message=f"فشل التحديث: {exc}",
            live_count=len(existing_live),
            archive_count=len(existing_archive),
        )
        
        return 1


if __name__ == "__main__":
    sys.exit(main())
